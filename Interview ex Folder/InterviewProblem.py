import zipfile  #to manage zipfile tools and compress end file

#testerfile = open(r'C:\Users\USERNAMEREMOVED\OneDrive\Desktop\test_results', 'r')
#with zipfile.ZipFile(testerfile, 'r') as my_zip:
#    print(my_zip.namelist())

# Permission denied. Run as Admin did not resolve. extracted files manually and moved to next step


# find file location
# convert file types to preferred
# update information in commonly used file type
# generate scatter plot used pre-set x and y-axis
# add line of best fit
# create image file
# remove excess data
# repeat scatter, Line of best fit, and create image


# to verify file path exists
import os.path
# Dataframe module - generate scatter plot
import pandas as pd
#Python to Excel read/write
from openpyxl import load_workbook
# Python to Excel, return Column letter
from openpyxl.utils import get_column_letter
# generate line of best fit (linear used)
import numpy

# TODO: Username remove, refactor path to intended source
#TODO: If recurring report:
# Pseudocode: on email receipt subject line content 'SUBJECT' if file exists download/convert/save

# Exception catch, if file_name.csv exists convert to xlsx format
if os.path.exists('Supplier_A_TestResults.csv'):
    SupAConvert = pd.read_csv('Supplier_A_TestResults.csv')
    SupAConvert.to_excel('Supplier_A_TestResults.xlsx', index=None, header=True)
    try:
        os.remove('Supplier_A_TestResults.csv')
    except FileNotFoundError:
        print('File not Found')
if os.path.exists('Supplier_B_TestResults.csv'):
    SupBConvert = pd.read_csv('Supplier_B_TestResults.csv')
    SupBConvert.to_excel('Supplier_B_TestResults.xlsx', index=None, header=True)
    try:
        os.remove('Supplier_B_TestResults.csv')
    except FileNotFoundError:
        print('File not Found')

# Exception catch, if file_name.json exists convert to xlsx format
if os.path.exists('Supplier_C_TestResults.json'):
    SupCConvert = pd.read_json('Supplier_C_TestResults.json')
    SupCConvert.to_excel('Supplier_C_TestResults.xlsx', index=None, header=True)
    try:
        os.remove('Supplier_C_TestResults.json')
    except FileNotFoundError:
        print('File not Found')


# Constants:
SupplierAFile = 'Supplier_A_TestResults.xlsx'
SupplierBFile = 'Supplier_B_TestResults.xlsx'
SupplierCFile = 'Supplier_C_TestResults.xlsx'



# TODO: add update to sheet name to 'Data Set' for (Force Value, Spring Displacement, Spring Constant)
# preset columns. assumed all future reports would be the same columns

def add_missing_values(filename):
    wb = load_workbook(filename)
    ws = wb.active
    # counter to identify last/max_row
    count = 0
    for row in ws:
        if not all([cell.value is None for cell in row]):
            count += 1
    # loop to insert header. #TODO: simplify and remove unnecessary loop
    for row in range(1, 2):
        for col in range(1, 8):
            char = get_column_letter(col)
            if row == 1 and char == 'E':
                ws[char + str(row)].value = 'Force Value'
            if row == 1 and char == 'F':
                ws[char + str(row)].value = 'Spring Displacement'
            if row == 1 and char == 'G':
                ws[char + str(row)].value = 'Spring Constant'
    # loop to generate formulas for calculating the above values
    for i in range(2, count + 1):
        ws['E' + str(i)].value = '=B' + str(i) + '*9.81'
        ws['F' + str(i)].value = '=D' + str(i) + '-C' + str(i)
        ws['G' + str(i)].value = '=E' + str(i) + '/F' + str(i)
    wb.save(filename)


# Generates scatter plot based on preset column names
def scatter(filename, image_name):
    df = pd.read_excel(filename)
    xplacer = df['Force Value']
    yplacer = df['Spring Displacement']
    pic = df.plot.scatter(title='Spring Displacement over Force', x=4,
                          y=5, grid=True, legend=False)
    # Line of best fit
    z = numpy.polyfit(xplacer, yplacer, 1)
    p = numpy.poly1d(z)
    pic.plot(xplacer, p(xplacer), "r--")
    # generate .jpg file for scatter plots
    pic = pic.get_figure()
    pic.savefig(str(image_name) + '.jpg')


# Remove results exceeding spring elastic limit (when starting measurement changes)
# TODO: update to remove one more result, as it can be assumed this is when the limit is exceeded
def remove_excess_scatter(filename, image_name, column='start_measurement_m'):
    df = pd.read_excel(filename)
    spring_threshold = df[column].min() + .01
    # Df of results less than spring_threshold
    df = df[df.start_measurement_m < spring_threshold]
    # TODO: Repeating code. Initial scatter function required filepath, could not accept updated df when inserted here
    xplacer = df['Force Value']
    yplacer = df['Spring Displacement']
    pic = df.plot.scatter(title='Spring Displacement over Force', x=4,
                          y=5, grid=True, legend=False)
    # line of best fit
    z = numpy.polyfit(xplacer, yplacer, 1)
    p = numpy.poly1d(z)
    pic.plot(xplacer, p(xplacer), "r--")
    # generate .jpg file for scatter plots
    pic = pic.get_figure()
    pic.savefig(str(image_name) + '.jpg')


#TODO: pandas reading excel formulas as NaN instead of returning the values
# Unable to find RereshAll to update formulas so pandas will return value > formula.
# Manual workaround, open each file, save and close file.
# solutions attempted time.sleep(5) to allow calculations to complete.
# pywin32 attempted but unsuccesful with current understanding
#TODO: See above. Run once to update files then comment to run a second time to finish the program.
# Function call to complete requested calculations across all 3 files, and update cells
add_missing_values(SupplierAFile)
add_missing_values(SupplierBFile)
add_missing_values(SupplierCFile)

# TODO: comment the above 3 lines and uncomment the following 6 lines to finish the program
#scatter(SupplierAFile, 'SupplierA_AllData_Scatter')
#scatter(SupplierBFile, 'SupplierB_AllData_Scatter')
#scatter(SupplierCFile, 'SupplierC_AllData_Scatter')
#remove_excess_scatter(SupplierAFile, image_name='SupplierA_Filtered_Scatter')
#remove_excess_scatter(SupplierBFile, image_name='SupplierB_Filtered_Scatter')
#remove_excess_scatter(SupplierCFile, image_name='SupplierC_Filtered_Scatter')
